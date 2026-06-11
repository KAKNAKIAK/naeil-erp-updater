"""Export helpers for V5 fare calculation results."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable, Mapping

from openpyxl import Workbook

from .calculator import RoundTripResult, js_round


def results_to_tsv(rows: Iterable[RoundTripResult]) -> str:
    lines = []
    for row in rows:
        value = "마감" if row.is_closed else str(js_round(row.total_fare))
        lines.append(f"{row.dep_date}\t{value}")
    return "\n".join(lines)


def export_results_to_excel(
    path: str | Path,
    results_by_night: Mapping[int, Iterable[RoundTripResult]],
    info: Mapping[str, object] | None = None,
) -> Path:
    path = Path(path)
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for night in sorted(results_by_night.keys()):
        ws = wb.create_sheet(f"{night}박")
        ws.append(["출발일", "왕복요금", "상태"])
        for row in results_by_night[night]:
            ws.append([
                row.dep_date,
                None if row.is_closed else js_round(row.total_fare),
                "마감" if row.is_closed else "",
            ])
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 10

    info_ws = wb.create_sheet("정보")
    info_ws.append(["항목", "값"])
    merged_info = {"생성일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
    if info:
        merged_info.update(dict(info))
    for key, value in merged_info.items():
        info_ws.append([key, "" if value is None else str(value)])
    info_ws.column_dimensions["A"].width = 18
    info_ws.column_dimensions["B"].width = 80

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def to_erp_rows(rows: Iterable[RoundTripResult], profit: int | None = None) -> list[dict[str, object]]:
    mapped = []
    for row in rows:
        if row.is_closed:
            continue
        mapped.append(
            {
                "date": row.dep_date,
                "adult_air": js_round(row.total_fare),
                "adult_hotel": "",
                "adult_land": "",
                "adult_tour": "",
                "adult_profit": "" if profit is None else int(profit or 0),
                "child_fare": "",
                "infant_fare": "",
            }
        )
    return mapped
